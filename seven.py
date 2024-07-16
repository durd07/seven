# -*- coding: utf-8 -*-
import datetime
import openpyxl
import numpy as np
import pandas as pd
import altair as alt
import streamlit as st
from io import BytesIO
from sklearn.preprocessing import minmax_scale

items_map = {
    '红细胞计数(RBC)(10^12/L)': ['红细胞计数', '红细胞'],
    '血红蛋白浓度(HGB)(g/L)': ['血红蛋白浓度', '血红蛋白'],
    '血小板计数(PLT)(10^9/L)': ['血小板计数', '血小板总数', '血小板'],
    '白细胞计数(WBC)(10^9/L)': ['白细胞计数', '白细胞数目', '白细胞'],
    '中性粒细胞绝对值(NEU#)(10^9/L)': ['中性粒细胞绝对值', '中性粒细胞计数', '嗜中性粒细胞绝对值'],
    '淋巴细胞绝对值(LYM#)(10^9/L)': ['淋巴细胞绝对值', '淋巴细胞计数'],
    '单核细胞绝对值(MON#)(10^9/L)': ['单核细胞绝对值', '单核细胞计数'],
    '嗜酸性粒细胞绝对值(EOS#)(10^9/L)': ['嗜酸性粒细胞绝对值', '嗜酸性粒细胞计数'],
    '嗜碱性粒细胞绝对值(BAS#)(10^9/L)': ['嗜碱性粒细胞绝对值', '嗜碱性粒细胞计数'],

    '中性粒细胞百分比(NEU%)(%)': ['中性粒细胞百分比', '嗜中性粒细胞百分比'],
    '淋巴细胞百分比(LYM%)(%)': ['淋巴细胞百分比'],
    '单核细胞百分比(MON%)(%)': ['单核细胞百分比'],
    '嗜酸性粒细胞百分比(EOS%)(%)': ['嗜酸性粒细胞百分比'],
    '嗜碱性粒细胞百分比(BAS%)(%)': ['嗜碱性粒细胞百分比'],

    '红细胞压积(HCT)(%)': ['红细胞压积'],
    '平均红细胞体积(MCV)(fL)': ['平均红细胞体积', '红细胞平均体积'],
    '平均红细胞血红蛋白含量(MCH)(pg)': ['平均红细胞血红蛋白含量', '平均血红蛋白含量', '平均红细胞血红蛋白'],
    '平均红细胞血红蛋白浓度(MCHC)(g/L)': ['平均红细胞血红蛋白浓度', '平均血红蛋白浓度'],
    '红细胞体积分布宽度-CV(RDW-CV)(%)': ['红细胞体积分布宽度-CV', '红细胞分布宽度CV', '红细胞分布宽度变异系数'],
    '红细胞体积分布宽度-SD(RDW-SD)(fL)': ['红细胞体积分布宽度-SD', '红细胞分布宽度-SD', '红细胞分布宽度标准差'],

    '血小板平均体积(MPV)(fL)': ['血小板平均体积', '平均血小板体积'],
    '大血小板比率(P-LC,R)': ['大血小板比率', '大型血小板比率'],
    '血小板压积(PCT)(%)': ['血小板压积'],
    '血小板体积分布宽度(PDW)(%)': ['血小板体积分布宽度', '血小板分布宽度'],

    # 'C-反应蛋白(CRP)(mg/L)', ['C-反应蛋白']
}


def construct_items_reverse_map(items_map):
    d = {}
    for k, v in items_map.items():
        for x in v:
            d[x] = k
    return d


dfs = {}
items_reverse_map = construct_items_reverse_map(items_map)


def load_data_with_new_format():
    dfs_new = pd.read_excel('杜子期血常规new.xlsx', engine='openpyxl',
                            sheet_name=None, index_col='编号', converters={'结果': float})

    dfs_local = {}
    for k, v in dfs_new.items():
        new_k = datetime.datetime.strptime(k, '%Y-%m-%d')
        v['项目'] = v['项目'].map(
            lambda x: items_reverse_map.get(x.split('(')[0], ''))
        v = v[v['项目'] != '']
        v = v.fillna('')
        v.index = np.arange(1, len(v) + 1)
        dfs_local[new_k.date()] = v
    return dfs_local


def load_data():
    df = pd.read_excel('杜子期血常规.xlsx', engine='openpyxl')

    dfs_local = {}
    for date, item in df.items():
        dfs_local.setdefault(date.date(), [])

        column_index = 1
        for i in range(0, len(item), 3):
            if item[i] is np.nan:
                continue

            name = item[i].split('(')[0]

            try:
                name = items_reverse_map[name]

                try:
                    value = float(item[i + 1])
                    min, max = item[i + 2].split('～')
                    if value < float(min):
                        stats = '↓'
                    elif value > float(max):
                        stats = '↑'
                    else:
                        stats = ''
                except:
                    value = np.nan
                    stats = ''

                dfs_local[date.date()].append({
                    '编号': column_index,
                    '项目': name,
                    '结果': value,
                    '状态': stats,
                    '参考值': item[i + 2],
                })
                column_index += 1
            except Exception as e:
                print(e)
    for date, v in dfs_local.items():
        df = pd.DataFrame.from_dict(v)
        df = df.set_index('编号')
        dfs_local[date] = df
    return dfs_local


def post_process():
    overall_df = pd.DataFrame()

    for k, v in sorted(dfs.items()):
        v = v.fillna('--')
        v = v.set_index('项目')
        overall_df[k] = v['结果']
    return overall_df


def columns_best_fit(ws: openpyxl.worksheet.worksheet.Worksheet):
        """
        Make all columns best fit
        """
        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].bestFit = True


def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=True, sheet_name='杜子期血常规数据统计')

    for k, v in dfs.items():
        v.to_excel(writer, index=True, sheet_name=str(k))

    workbook = writer.book
    #format1 = workbook.add_format({'num_format': '0.00'})
    for _, sheet in writer.sheets.items():
        #sheet.set_column('A:Z', None, format1)

        #from openpyxl.utils import get_column_letter
        #for column_index in range(1, 6):
        #    excel_column_name = get_column_letter(column_index)
        #    st.write(excel_column_name)
        #    sheet.column_dimensions[excel_column_name].bestFit = True
        #    #sheet.column_dimensions[excel_column_name].auto_fit = True

        columns_best_fit(sheet)
    writer.close()
    processed_data = output.getvalue()
    return processed_data

def to_excel2(dfs):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    k_sorted = list(dfs.keys())
    k_sorted.sort()
    for x in k_sorted:
        dfs[x].to_excel(writer, index=True, sheet_name=str(x))

    workbook = writer.book
    #format1 = workbook.add_format({'num_format': '0.00'})
    for _, sheet in writer.sheets.items():
        #sheet.set_column('A:Z', None, format1)

        #from openpyxl.utils import get_column_letter
        #for column_index in range(1, 6):
        #    excel_column_name = get_column_letter(column_index)
        #    st.write(excel_column_name)
        #    sheet.column_dimensions[excel_column_name].bestFit = True
        #    #sheet.column_dimensions[excel_column_name].auto_fit = True

        columns_best_fit(sheet)
    writer.close()
    processed_data = output.getvalue()
    return processed_data


def display(df):
    def highlight_dataframe(s):
        lst = []
        date = s.name
        for k, v in s.items():
            try:
                stats = dfs[date].loc[dfs[date]['项目'] == k]['状态'].item()
                if stats == '↑':
                    lst.append('color: red')
                elif stats == '↓':
                    lst.append('color: orange')
                else:
                    lst.append('')
            except Exception as e:
                lst.append('')
        return lst

    df_str = df.astype(str)

    st.set_page_config(layout='wide')
    pd.set_option("display.max_colwidth", 1000, 'display.width', 1000)

    st.title('杜子期血常规数据统计')
    st.write(df_str.style.apply(highlight_dataframe, axis=0))
    st.download_button("Export to Excel", data=to_excel(df),
                       file_name='杜子期血常规数据统计.xlsx')

    all_draw_items = items_map.keys()

    chart_items = list()
    st.sidebar.write('请选择画图项')
    if st.sidebar.checkbox('所有项'):
        chart_items = all_draw_items

    for item in all_draw_items:
        if st.sidebar.checkbox(item):
            chart_items.append(item)

    if chart_items:
        df_chart = df.loc[chart_items, :].T
        #st.line_chart(df)
    else:
        df_chart = df.loc[['血小板计数(PLT)(10^9/L)'], :].T
        #st.line_chart(df)

    df_chart['date'] = df_chart.index

    options = np.array(df_chart['date']).tolist()

    (start_time, end_time) = st.select_slider("**请选择时间序列长度：**",
         #min_value = datetime(2013, 10, 1,),
         #max_value = datetime(2018, 10, 31,),
         options = options,
         value= (options[0],options[-1],),
     )

    df_chart = df_chart[(df_chart['date']>=start_time) & (df_chart['date']<=end_time)]

    with st.expander("**点击展开查看单次详细信息**"):
        tabs = st.tabs([str(x) for x in df_chart.index])
        for i in range(0, len(tabs)):
            with tabs[i]:
                st.dataframe(dfs[df_chart.index[i]].astype(str), use_container_width=True, height = min((dfs[df_chart.index[i]].shape[0] + 1) * 35 + 3, 17*35+3),)

        st.download_button("Export to Excel", data=to_excel2(dfs),
                           file_name='杜子期血常规数据统计2.xlsx')


    df_chart = df_chart.replace('--', np.nan)

    st.write("### 图表展示")
    st.write("#### 单表展示")
    for column in df_chart.columns:
        if column == 'date':
            continue

        min_v = df_chart[column].min()
        max_v = df_chart[column].max()
        scale_min = max(min_v - (max_v - min_v) * 0.1, 0)
        scale_max = max_v + (max_v - min_v) * 0.1

        st.write(f"**{column}**")
        st.vega_lite_chart(data=df_chart, spec={
            'layer': [
                {
                    'mark': {
                        'type': 'line',
                        'point': {"filled": False, "fill": "white"},
                        'tooltip': True,
                        'strokeWidth': 3
                    }
                },
                {
                    'mark': {
                        'type': 'text',
                        'align': 'center',
                        'baseline': 'line-bottom',
                        'dx': 3,
                        'fontSize': 20,
                        'fontWeight': 'normal'
                    },
                    'encoding': {
                        'text': {'field': column, 'type': 'quantitative'}
                    }
                }
            ],
            'encoding': {
                'x': {
                    "type": "temporal",
                    #'timeUnit': 'date',
                    'title': "时间",
                    'field': 'date',
                    "axis": {
                        "format": "%y-%m-%d",
						"labelAngle": -30,
						"labelColor": 'black',
						"labelFontSize": 18,
                        'titleColor': 'black',
                        }
                    },
                'y': {
                    "type": "quantitative",
                    'field': column,
                    "scale": {"domain": [scale_min, scale_max]},
                    #'aggregate': 'mean'
                    "axis": {
						"labelColor": 'black',
						"labelFontSize": 18,
                        'titleColor': 'black',
                    },
                    },
                #'color': {'field': 'field', 'type': 'nominal'},
                },
            }, use_container_width=True)

    st.write("#### 合并展示")
    df_chart = df_chart.drop('date', axis=1)
    df_chart[df_chart.columns] = minmax_scale(df_chart)
    ndf = df_chart.melt(var_name='field', value_name='data')
    xx = pd.concat([df_chart.index.to_series()] * int((ndf.shape[0] / len(df_chart.index))))
    ndf['date'] = xx.values

    st.vega_lite_chart(data=ndf, spec={
        'layer': [
            {
                'mark': {
                    'type': 'line',
                    'point': {"filled": False, "fill": "white"},
                    'tooltip': True,
                    'strokeWidth': 3
                }
            },
    #        {
    #            'mark': {
    #                'type': 'text',
    #                'align': 'center',
    #                'baseline': 'line-bottom',
    #                'dx': 3,
    #                'size': 14
    #            },
    #            'encoding': {
    #                'text': {'field': 'data', 'type': 'quantitative'}
    #            }
    #        }
        ],
        'encoding': {
            'x': {
                    "type": "temporal",
                    #'timeUnit': 'date',
                    'title': "时间",
                    'field': 'date',
                    "axis": {
                        "format": "%y-%m-%d",
						"labelAngle": -30,
						"labelColor": 'black',
						"labelFontSize": '18',
                        'titleColor': 'black',
                        }
                },
            'y': {
                "type": "quantitative",
                #'field': '血小板计数(PLT)(10^9/L)'
                'field': 'data',
                #'aggregate': 'mean'
                    #'aggregate': 'mean'
                    "axis": {
						"labelColor": 'black',
						"labelFontSize": '18',
                        'titleColor': 'black',
                    },
                    },
                #'color': {'field': 'field', 'type': 'nominal'},
            'color': {'field': 'field', 'type': 'nominal'},
            },
        }, use_container_width=True)

    #st.write('### 相关系数矩阵')
    #df = df.filter(regex='^((?!_参考范围$).)*$', axis=0).astype(float)
    #df = df.filter(regex='^((?!_参考范围$).)*$', axis=0).astype(float)
    df = df.replace('--', np.nan)
    df.index.name = None
    #st.write(df.T.corr())


    cor_data = df.T.corr().stack().reset_index().rename(columns={0: 'correlation', 'level_0': 'variable', 'level_1': 'variable2'})
    cor_data['correlation_label'] = cor_data['correlation'].map('{:.2f}'.format)

    base = alt.Chart(cor_data).encode(
        x='variable2:O',
        y='variable:O'
    )

    # Text layer with correlation labels
    # Colors are for easier readability
    text = base.mark_text().encode(
        text='correlation_label',
        color=alt.condition(
            alt.datum.correlation > 0.5,
            alt.value('white'),
            alt.value('black')
        )
    )

    # The correlation heatmap itself
    cor_plot = base.mark_rect().encode(
        color='correlation:Q'
    )

    st.write("### 相关系数热力图")
    st.altair_chart(cor_plot + text, use_container_width=True)


def run():
    dfs.update(load_data())
    dfs.update(load_data_with_new_format())
    overall_df = post_process()

    overall_df = overall_df.fillna('--')
    overall_df = overall_df.reindex(items_map)
    display(overall_df)


    df = pd.read_excel('杜子期血常规.xlsx', engine='openpyxl', sheet_name='schedule')
    st.dataframe(df, hide_index=True)

run()
